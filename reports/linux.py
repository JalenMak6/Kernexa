"""
reports/linux.py
Linux kernel compliance report:
  - build_xlsx()        — builds Excel workbook from scan data
  - send_scan_report()  — emails the report with xlsx attachment
"""

import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formatdate

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reports.common import classify_os, sheet_sort_key, xl_header_style, xl_cell_border
from database import get_notification_settings


def build_xlsx(scan_data: dict) -> bytes:
    """Build Linux kernel compliance Excel workbook from scan data."""
    HEADERS    = [
        "Host", "OS", "Kernel Status", "Current Kernel", "Latest Kernel",
        "Pending Pkg Count", "Pending Security Package", "Last Reboot", "Advisory Count",
    ]
    COL_WIDTHS = [38, 20, 14, 32, 32, 16, 40, 20, 14]

    GREEN_FILL  = PatternFill("solid", fgColor="F0FDF4")
    RED_FILL    = PatternFill("solid", fgColor="FEF2F2")
    NORMAL_FONT = Font(size=10)

    hdr_fill, hdr_font, hdr_align, hdr_border = xl_header_style()
    cell_border = xl_cell_border()

    def write_sheet(ws, hosts):
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20
        for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = hdr_align
            cell.border    = hdr_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        row_idx = 2
        for host in hosts:
            current      = host.get("current_kernel_version", "") or ""
            latest       = host.get("latest_available_kernel_version", "") or ""
            compliant    = bool(current and latest and current == latest)
            pkgs         = host.get("pending_security_packages", []) or []
            pkg_count    = len(pkgs)
            status       = "Up to date" if compliant else "Outdated"
            last_reboot  = host.get("last_reboot_time", "") or ""
            advisory_cnt = len(host.get("advisory_ids", []) or [])

            status_fill  = GREEN_FILL if compliant else RED_FILL
            status_color = "166534"   if compliant else "991B1B"
            pkg_rows     = pkgs if pkgs else [""]
            first_row    = row_idx

            for pkg_i, pkg in enumerate(pkg_rows):
                is_first = (pkg_i == 0)
                values   = [
                    host.get("host", ""),
                    host.get("os_version", ""),
                    status, current, latest, pkg_count, pkg, last_reboot, advisory_cnt,
                ]
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border    = cell_border
                    cell.alignment = Alignment(vertical="center")
                    cell.font      = NORMAL_FONT
                    if col_idx == 3 and is_first:
                        cell.fill = status_fill
                        cell.font = Font(bold=True, color=status_color, size=10)
                    elif col_idx == 6 and is_first and pkg_count > 0:
                        cell.font = Font(bold=True, color="C2410C", size=10)
                    elif col_idx == 7 and pkg:
                        cell.font = Font(name="Courier New", size=9)
                ws.row_dimensions[row_idx].height = 15
                row_idx += 1

            if row_idx > first_row:
                thick = Side(style="medium", color="CBD5E1")
                for col_idx in range(1, len(HEADERS) + 1):
                    cell = ws.cell(row=row_idx - 1, column=col_idx)
                    cell.border = Border(
                        left=cell.border.left, right=cell.border.right,
                        top=cell.border.top, bottom=thick,
                    )

    hosts  = scan_data.get("hosts", [])
    groups: dict[str, list] = {}
    for h in hosts:
        key = classify_os(h.get("os_version", ""))
        groups.setdefault(key, []).append(h)

    sorted_keys = sorted(groups.keys(), key=sheet_sort_key)
    wb = openpyxl.Workbook()
    ws_summary = wb.active
    ws_summary.title = "All Hosts"
    write_sheet(ws_summary, hosts)
    for key in sorted_keys:
        ws = wb.create_sheet(title=key[:31])
        write_sheet(ws, groups[key])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_scan_report(scan_data: dict, scan_id: str):
    """Email the Linux kernel compliance report as an xlsx attachment."""
    try:
        settings = get_notification_settings()
        if not settings["smtp_host"] or not settings["recipients"]:
            print("Email notifications: no SMTP host or recipients configured, skipping")
            return

        xlsx_bytes = build_xlsx(scan_data)
        scanned_at = scan_data.get("scanned_at", "")
        host_count = len(scan_data.get("hosts", []))
        compliant  = sum(
            1 for h in scan_data.get("hosts", [])
            if h.get("current_kernel_version") == h.get("latest_available_kernel_version")
            and h.get("current_kernel_version")
        )
        pct = round((compliant / host_count) * 100) if host_count else 0

        msg = MIMEMultipart()
        msg["From"]    = settings["smtp_from"] or settings["smtp_user"]
        msg["To"]      = ", ".join(settings["recipients"])
        msg["Date"]    = formatdate(localtime=True)
        msg["Subject"] = f"Kernexa Scan Report — {host_count} hosts, {pct}% compliant"

        body = (
            f"Kernexa Linux scan completed.\n\n"
            f"  Scan ID:    {scan_id}\n"
            f"  Scanned at: {scanned_at}\n"
            f"  Hosts:      {host_count}\n"
            f"  Compliant:  {compliant} ({pct}%)\n"
            f"  Outdated:   {host_count - compliant}\n\n"
            f"Full results attached as an Excel workbook (.xlsx) with one sheet per OS group.\n"
        )
        msg.attach(MIMEText(body, "plain"))

        attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        attachment.set_payload(xlsx_bytes)
        encoders.encode_base64(attachment)
        attachment.add_header(
            "Content-Disposition",
            f'attachment; filename="kernexa-scan-{scan_id[:8]}.xlsx"'
        )
        msg.attach(attachment)

        if settings["tls_enabled"]:
            server = smtplib.SMTP(settings["smtp_host"], settings["smtp_port"], timeout=15)
            server.ehlo(); server.starttls(); server.ehlo()
        else:
            server = smtplib.SMTP_SSL(settings["smtp_host"], settings["smtp_port"], timeout=15)

        if settings["smtp_user"] and settings["smtp_password"]:
            server.login(settings["smtp_user"], settings["smtp_password"])

        server.sendmail(msg["From"], settings["recipients"], msg.as_string())
        server.quit()
        print(f"Linux scan report emailed to {settings['recipients']}")

    except Exception as e:
        print(f"send_scan_report error: {e}")