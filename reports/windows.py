"""
reports/windows.py
Windows patch compliance report:
  - build_windows_xlsx()        — builds Excel workbook from KB records
  - send_windows_scan_report()  — emails the report with xlsx attachment
"""

import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formatdate

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reports.common import classify_os, sheet_sort_key, xl_header_style, xl_cell_border
from database import get_notification_settings


def build_windows_xlsx(records: list) -> bytes:
    """Build Windows patch compliance Excel workbook — one row per KB per host."""
    HEADERS    = [
        "Hostname", "OS Name", "OS Version", "KBID", "Patch Name",
        "Classification", "MSRC Severity", "Reboot Required", "Published Date",
    ]
    COL_WIDTHS = [38, 32, 16, 14, 60, 22, 18, 20, 22]

    hdr_fill, hdr_font, hdr_align, hdr_border = xl_header_style()
    cell_border = xl_cell_border()
    NORMAL_FONT = Font(size=10)

    SEV_FILL = {
        "Critical":  PatternFill("solid", fgColor="FEF2F2"),
        "Important": PatternFill("solid", fgColor="FFF7ED"),
        "Moderate":  PatternFill("solid", fgColor="FEFCE8"),
        "Low":       PatternFill("solid", fgColor="F0FDF4"),
    }
    SEV_COLOR = {
        "Critical":  "991B1B",
        "Important": "9A3412",
        "Moderate":  "854D0E",
        "Low":       "166534",
    }

    def write_sheet(ws, rows):
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20
        for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = hdr_align
            cell.border    = hdr_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        for row_idx, r in enumerate(rows, start=2):
            sev = r.get("msrcSeverity", "") or ""
            pub = r.get("publishedDateTime", "") or ""
            if pub and "T" in pub:
                try:
                    pub = pub[:10]
                except Exception:
                    pass

            values = [
                r.get("hostname", ""),
                r.get("osName", ""),
                r.get("osVersion", ""),
                r.get("KBID", ""),
                r.get("patchName", ""),
                r.get("classification", ""),
                sev,
                r.get("rebootRequired", ""),
                pub,
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border    = cell_border
                cell.alignment = Alignment(vertical="center")
                cell.font      = NORMAL_FONT
                if col_idx == 7 and sev in SEV_FILL:
                    cell.fill = SEV_FILL[sev]
                    cell.font = Font(bold=True, color=SEV_COLOR[sev], size=10)
            ws.row_dimensions[row_idx].height = 15

    groups: dict[str, list] = {}
    for r in records:
        key = classify_os(r.get("osName", ""))
        groups.setdefault(key, []).append(r)

    sorted_keys = sorted(groups.keys(), key=sheet_sort_key)
    wb = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "All Hosts"
    write_sheet(ws_all, records)
    for key in sorted_keys:
        ws = wb.create_sheet(title=key[:31])
        write_sheet(ws, groups[key])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_windows_scan_report(records: list, scan_id: str, scanned_at: str):
    """Email the Windows patch compliance report as an xlsx attachment."""
    try:
        settings = get_notification_settings()
        if not settings["smtp_host"] or not settings["recipients"]:
            print("Email notifications: no SMTP host or recipients configured, skipping")
            return

        host_set     = set(r["hostname"] for r in records)
        host_count   = len(host_set)
        total_kbs    = len(records)
        sec_hosts    = len({r["hostname"] for r in records
                            if r.get("classification") in ("Security Updates", "Critical Updates")})
        reboot_hosts = len({r["hostname"] for r in records
                            if r.get("rebootRequired") in ("AlwaysRequiresReboot", "CanRequestReboot")})

        xlsx_bytes = build_windows_xlsx(records)

        msg = MIMEMultipart()
        msg["From"]    = settings["smtp_from"] or settings["smtp_user"]
        msg["To"]      = ", ".join(settings["recipients"])
        msg["Date"]    = formatdate(localtime=True)
        msg["Subject"] = f"Kernexa Windows Scan Report — {host_count} hosts, {total_kbs} pending KBs"

        host_map: dict[str, list] = {}
        for r in records:
            host_map.setdefault(r["hostname"], []).append(r)

        host_lines = []
        for hostname in sorted(host_map.keys()):
            kbs     = host_map[hostname]
            sec     = sum(1 for k in kbs if k.get("classification") in ("Security Updates", "Critical Updates"))
            ru      = sum(1 for k in kbs if k.get("classification") == "Update Rollups")
            defs    = sum(1 for k in kbs if k.get("classification") == "Definition Updates")
            reboot  = any(k.get("rebootRequired") in ("AlwaysRequiresReboot", "CanRequestReboot") for k in kbs)
            os_name = kbs[0].get("osName", "").replace("Microsoft Windows Server", "WS").replace("Microsoft Windows", "Win")

            parts  = []
            if sec:  parts.append(f"{sec} Security")
            if ru:   parts.append(f"{ru} Rollup")
            if defs: parts.append(f"{defs} Definition")
            kb_ids = ", ".join(f"KB{k['KBID']}" for k in kbs[:5])
            if len(kbs) > 5:
                kb_ids += f" (+{len(kbs)-5} more)"

            host_lines.append(
                f"  {hostname}\n"
                f"    OS: {os_name}  |  KBs: {', '.join(parts)}  |  Reboot: {'Yes' if reboot else 'No'}\n"
                f"    Patches: {kb_ids}"
            )

        body = (
            f"Kernexa Windows patch compliance scan completed.\n\n"
            f"  Scan ID:           {scan_id}\n"
            f"  Scanned at:        {scanned_at}\n"
            f"  Windows Hosts:     {host_count}\n"
            f"  Total Pending KBs: {total_kbs}\n"
            f"  Hosts w/ Security: {sec_hosts}\n"
            f"  Hosts need Reboot: {reboot_hosts}\n\n"
            f"{'='*60}\n"
            f"HOST SUMMARY\n"
            f"{'='*60}\n"
            + "\n\n".join(host_lines) +
            f"\n\n{'='*60}\n"
            f"Full results attached as Excel workbook (.xlsx) with one sheet per OS version.\n"
        )
        msg.attach(MIMEText(body, "plain"))

        attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        attachment.set_payload(xlsx_bytes)
        encoders.encode_base64(attachment)
        attachment.add_header(
            "Content-Disposition",
            f'attachment; filename="kernexa-windows-{scan_id[:8]}.xlsx"'
        )
        msg.attach(attachment)

        if settings["tls_enabled"]:
            server = smtplib.SMTP(settings["smtp_host"], settings["smtp_port"], timeout=15)
            server.ehlo(); server.starttls(); server.ehlo()
        else:
            server = smtplib.SMTP_SSL(settings["smtp_host"], settings["smtp_port"], timeout=15)

        if settings["smtp_user"] and settings["smtp_password"]:
            server.login(settings["smtp_user"], settings["smtp_password"])

        server.sendmail(msg["From"], settings["recipients"], msg.as_string())
        server.quit()
        print(f"Windows scan report emailed to {settings['recipients']}")

    except Exception as e:
        print(f"send_windows_scan_report error: {e}")