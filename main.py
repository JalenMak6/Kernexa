from fastapi import FastAPI, BackgroundTasks, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from apscheduler.schedulers.background import BackgroundScheduler
from contextlib import asynccontextmanager
from datetime import datetime, timezone
from pydantic import BaseModel
from typing import List, Optional
import uuid
import os
import threading
import re
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from email.utils import formatdate

from scanner import run_patch_scan
from enricher import enrich_all
from database import (
    save_to_db, get_latest_scan, get_scan_history,
    get_inventories, save_inventory, activate_inventory,
    delete_inventory, get_inventory_content,
    save_credentials, get_credentials, get_active_inventory_credentials,
    get_tags_for_host, get_all_tags, add_tag, remove_tag,
    get_cve_details, get_scan_failures, get_host_history, get_host_cves,
    get_notification_settings, save_notification_settings
)


INVENTORY_PATH = "./inventory/hosts"

# ── models ────────────────────────────────────────────────────────────────────

class HostsUpdate(BaseModel):
    hosts: List[str]

class TagAdd(BaseModel):
    tag: str

class CredentialsUpdate(BaseModel):
    inventory_id: int
    username: str
    password: str

class NotificationSettings(BaseModel):
    smtp_host: str
    smtp_port: int = 587
    smtp_user: str = ""
    smtp_password: str = ""
    smtp_from: str = ""
    recipients: List[str] = []
    tls_enabled: bool = True

# ── helpers ───────────────────────────────────────────────────────────────────

def parse_hosts_from_content(content: str) -> list:
    return [
        l.strip() for l in content.splitlines()
        if l.strip() and not l.startswith("[") and not l.startswith("#")
    ]

def write_inventory_content(content: str):
    os.makedirs(os.path.dirname(INVENTORY_PATH), exist_ok=True)
    with open(INVENTORY_PATH, "w") as f:
        f.write(content)

running_scans = {}
_scan_lock = threading.Lock()

# ── email helpers ─────────────────────────────────────────────────────────────

def classify_os(os_version: str) -> str:
    """Map a full OS version string to a short sheet name."""
    if not os_version:
        return "Unknown"
    v = os_version.lower()
    # RHEL / Red Hat
    if "red hat" in v or "rhel" in v:
        m = re.search(r'(\d+)', os_version)
        return f"RHEL {m.group(1)}" if m else "RHEL"
    # Rocky Linux
    if "rocky" in v:
        m = re.search(r'(\d+)', os_version)
        return f"Rocky {m.group(1)}" if m else "Rocky"
    # AlmaLinux
    if "alma" in v:
        m = re.search(r'(\d+)', os_version)
        return f"Alma {m.group(1)}" if m else "Alma"
    # CentOS
    if "centos" in v:
        m = re.search(r'(\d+)', os_version)
        return f"CentOS {m.group(1)}" if m else "CentOS"
    # Ubuntu
    if "ubuntu" in v:
        m = re.search(r'(\d+\.\d+)', os_version)
        return f"Ubuntu {m.group(1)}" if m else "Ubuntu"
    # Debian
    if "debian" in v:
        m = re.search(r'(\d+)', os_version)
        return f"Debian {m.group(1)}" if m else "Debian"
    return os_version[:31]  # Excel sheet name limit


def _xl_header_style():
    """Return header fill, font, alignment, border."""
    fill   = PatternFill("solid", fgColor="0F172A")
    font   = Font(bold=True, color="FFFFFF", size=10)
    align  = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin   = Side(style="thin", color="334155")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return fill, font, align, border


def _xl_cell_border():
    thin = Side(style="thin", color="E2E8F0")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def build_xlsx(scan_data: dict) -> bytes:
    """
    Build an Excel workbook with:
      - One summary sheet (all hosts)
      - One sheet per OS group (RHEL 7, RHEL 8, Ubuntu 22.04, etc.)
    Each sheet has: Host | OS | Kernel Status | Current Kernel | Latest Kernel |
                    Pending Pkg Count | Pending Packages | Last Reboot | Advisory Count
    """
    import re

    HEADERS = [
        "Host", "OS", "Kernel Status", "Current Kernel", "Latest Kernel",
        "Pending Pkg Count", "Pending Security Package", "Last Reboot", "Advisory Count",
    ]
    COL_WIDTHS = [38, 20, 14, 32, 32, 16, 40, 20, 14]

    # Status colours
    GREEN_FILL  = PatternFill("solid", fgColor="F0FDF4")
    RED_FILL    = PatternFill("solid", fgColor="FEF2F2")
    ALT_FILL    = PatternFill("solid", fgColor="FAFAFA")   # alternating continuation rows
    NORMAL_FONT = Font(size=10)

    hdr_fill, hdr_font, hdr_align, hdr_border = _xl_header_style()
    cell_border = _xl_cell_border()

    def write_sheet(ws, hosts):
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 20

        # Headers
        for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill      = hdr_fill
            cell.font      = hdr_font
            cell.alignment = hdr_align
            cell.border    = hdr_border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        row_idx = 2
        for host in hosts:
            current   = host.get("current_kernel_version", "") or ""
            latest    = host.get("latest_available_kernel_version", "") or ""
            compliant = bool(current and latest and current == latest)
            pkgs      = host.get("pending_security_packages", []) or []
            pkg_count = len(pkgs)
            status    = "Up to date" if compliant else "Outdated"
            last_reboot  = host.get("last_reboot_time", "") or ""
            advisory_cnt = len(host.get("advisory_ids", []) or [])

            status_fill = GREEN_FILL if compliant else RED_FILL
            status_color = "166534" if compliant else "991B1B"

            # One row per package; if no packages, still write one row
            pkg_rows = pkgs if pkgs else [""]
            first_row = row_idx

            for pkg_i, pkg in enumerate(pkg_rows):
                is_first = (pkg_i == 0)
                row_fill = status_fill if is_first else ALT_FILL

                values = [
                    host.get("host", ""),
                    host.get("os_version", ""),
                    status,
                    current,
                    latest,
                    pkg_count,
                    pkg,
                    last_reboot,
                    advisory_cnt,
                ]

                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border    = cell_border
                    cell.alignment = Alignment(vertical="center")
                    cell.font      = NORMAL_FONT

                    # Status cell styling
                    if col_idx == 3 and is_first:
                        cell.fill = status_fill
                        cell.font = Font(bold=True, color=status_color, size=10)
                    # Pkg count cell — orange if > 0
                    elif col_idx == 6 and is_first and pkg_count > 0:
                        cell.font = Font(bold=True, color="C2410C", size=10)
                    # Package name — monospace
                    elif col_idx == 7 and pkg:
                        cell.font = Font(name="Courier New", size=9)

                ws.row_dimensions[row_idx].height = 15
                row_idx += 1

            # Draw a thicker bottom border on the last row of this host group
            if row_idx > first_row:
                thick = Side(style="medium", color="CBD5E1")
                for col_idx in range(1, len(HEADERS) + 1):
                    cell = ws.cell(row=row_idx - 1, column=col_idx)
                    cell.border = Border(
                        left=cell.border.left,
                        right=cell.border.right,
                        top=cell.border.top,
                        bottom=thick,
                    )

    # ── Group hosts by OS ────────────────────────────────────────────────────
    hosts = scan_data.get("hosts", [])
    groups: dict[str, list] = {}
    for h in hosts:
        key = classify_os(h.get("os_version", ""))
        groups.setdefault(key, []).append(h)

    # Sort sheet names naturally: RHEL 7 < RHEL 8 < Rocky 8 < Ubuntu 22.04 …
    def sheet_sort_key(name):
        parts = name.split()
        prefix = parts[0] if parts else ""
        num_str = parts[1] if len(parts) > 1 else "0"
        try:
            num = float(num_str)
        except ValueError:
            num = 0
        return (prefix, num)

    sorted_keys = sorted(groups.keys(), key=sheet_sort_key)

    wb = openpyxl.Workbook()

    # Summary sheet first
    ws_summary = wb.active
    ws_summary.title = "All Hosts"
    write_sheet(ws_summary, hosts)

    # Per-OS sheets
    for key in sorted_keys:
        ws = wb.create_sheet(title=key[:31])
        write_sheet(ws, groups[key])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def send_scan_report(scan_data: dict, scan_id: str):
    try:
        settings = get_notification_settings()
        if not settings["smtp_host"] or not settings["recipients"]:
            print("Email notifications: no SMTP host or recipients configured, skipping")
            return

        xlsx_bytes  = build_xlsx(scan_data)
        scanned_at  = scan_data.get("scanned_at", "")
        host_count  = len(scan_data.get("hosts", []))
        compliant   = sum(
            1 for h in scan_data.get("hosts", [])
            if h.get("current_kernel_version") == h.get("latest_available_kernel_version")
            and h.get("current_kernel_version")
        )
        pct = round((compliant / host_count) * 100) if host_count else 0

        msg = MIMEMultipart()
        msg["From"]    = settings["smtp_from"] or settings["smtp_user"]
        msg["To"]      = ", ".join(settings["recipients"])
        msg["Date"]    = formatdate(localtime=True)
        msg["Subject"] = f"Kernexa Scan Report — {host_count} hosts, {pct}% compliant"

        body = (
            f"Kernexa scan completed.\n\n"
            f"  Scan ID:    {scan_id}\n"
            f"  Scanned at: {scanned_at}\n"
            f"  Hosts:      {host_count}\n"
            f"  Compliant:  {compliant} ({pct}%)\n"
            f"  Outdated:   {host_count - compliant}\n\n"
            f"Full results are attached as an Excel workbook (.xlsx), "
            f"with one sheet per OS group.\n"
        )
        msg.attach(MIMEText(body, "plain"))

        attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        attachment.set_payload(xlsx_bytes)
        encoders.encode_base64(attachment)
        attachment.add_header(
            "Content-Disposition",
            f'attachment; filename="kernexa-scan-{scan_id[:8]}.xlsx"'
        )
        msg.attach(attachment)

        if settings["tls_enabled"]:
            server = smtplib.SMTP(settings["smtp_host"], settings["smtp_port"], timeout=15)
            server.ehlo()
            server.starttls()
            server.ehlo()
        else:
            server = smtplib.SMTP_SSL(settings["smtp_host"], settings["smtp_port"], timeout=15)

        if settings["smtp_user"] and settings["smtp_password"]:
            server.login(settings["smtp_user"], settings["smtp_password"])

        server.sendmail(msg["From"], settings["recipients"], msg.as_string())
        server.quit()
        print(f"Scan report emailed to {settings['recipients']}")

    except Exception as e:
        print(f"send_scan_report error: {e}")

def run_and_save(scan_id: str, scanned_at: datetime):
    try:
        running_scans[scan_id] = "running"
        output = run_patch_scan()
        save_to_db(output, scan_id, scanned_at)
        running_scans[scan_id] = "enriching"
        print(f"Scan {scan_id} saved — starting CVE enrichment")
        enrich_all()
        print(f"Scan {scan_id} enrichment complete")
        running_scans[scan_id] = "complete"
        try:
            scan_data = get_latest_scan()
            if scan_data:
                send_scan_report(scan_data, scan_id)
        except Exception as e:
            print(f"Email report error (non-fatal): {e}")
    except Exception as e:
        running_scans[scan_id] = f"failed: {str(e)}"
        print(f"run_and_save error: {e}")

def scheduled_scan():
    """Called by APScheduler every 10 minutes. Skips if a scan is already running."""
    with _scan_lock:
        already_running = any(
            v in ("running", "pending")
            for v in running_scans.values()
        )
        if already_running:
            print("Scheduled scan skipped: a scan is already in progress")
            return

        creds = get_active_inventory_credentials()
        if not creds:
            print("Scheduled scan skipped: no credentials set for active inventory")
            return

        scan_id    = str(uuid.uuid4())
        scanned_at = datetime.now(timezone.utc)
        running_scans[scan_id] = "pending"

    print(f"Scheduled scan starting: {scan_id}")
    run_and_save(scan_id, scanned_at)

# ── scheduler ─────────────────────────────────────────────────────────────────

scheduler = BackgroundScheduler()

# ── startup ───────────────────────────────────────────────────────────────────

@asynccontextmanager
async def lifespan(app: FastAPI):
    try:
        inventories = get_inventories()
        if not inventories:
            if os.path.exists(INVENTORY_PATH):
                with open(INVENTORY_PATH, "r") as f:
                    content = f.read()
                hosts = parse_hosts_from_content(content)
                if hosts:
                    inv_id = save_inventory("Default Inventory", content, len(hosts))
                    activate_inventory(inv_id)
                    print(f"Seeded default inventory with {len(hosts)} hosts")
            else:
                print("No default inventory file found at ./inventory/hosts")
        else:
            active = next((i for i in inventories if i["is_active"]), None)
            if active:
                content = get_inventory_content(active["id"])
                write_inventory_content(content)
                print(f"Restored active inventory '{active['name']}' with {active['host_count']} hosts")
    except Exception as e:
        print(f"Startup error: {e}")

    scheduler.add_job(scheduled_scan, "interval", minutes=10, id="auto_scan")
    scheduler.start()
    print("Auto-scan scheduler started (every 10 minutes)")

    yield

    scheduler.shutdown(wait=False)
    print("Scheduler stopped")

# ── app ───────────────────────────────────────────────────────────────────────

_docs_enabled = os.getenv("ENABLE_DOCS", "false").lower() == "true"
app = FastAPI(
    title="Patch Scan Platform",
    lifespan=lifespan,
    docs_url="/docs"         if _docs_enabled else None,
    redoc_url="/redoc"       if _docs_enabled else None,
    openapi_url="/openapi.json" if _docs_enabled else None,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── inventory endpoints ───────────────────────────────────────────────────────

@app.get("/api/inventories")
async def list_inventories():
    return get_inventories()

@app.post("/api/inventories/upload")
async def upload_inventory(file: UploadFile = File(...), name: str = Form(...)):
    content = (await file.read()).decode("utf-8")
    hosts = parse_hosts_from_content(content)
    if not hosts:
        raise HTTPException(status_code=400, detail="No valid hosts found in file")
    inv_id = save_inventory(name, content, len(hosts))
    return {"id": inv_id, "name": name, "host_count": len(hosts), "message": f"Uploaded {len(hosts)} hosts"}

@app.post("/api/inventories/{inv_id}/activate")
async def set_active_inventory(inv_id: int):
    content = activate_inventory(inv_id)
    write_inventory_content(content)
    hosts = parse_hosts_from_content(content)
    return {"message": f"Activated inventory with {len(hosts)} hosts"}

@app.delete("/api/inventories/{inv_id}")
async def remove_inventory(inv_id: int):
    delete_inventory(inv_id)
    return {"message": "Deleted"}

# ── credential endpoints ──────────────────────────────────────────────────────

@app.post("/api/credentials")
async def set_credentials(body: CredentialsUpdate):
    if not body.username or not body.password:
        raise HTTPException(status_code=400, detail="Username and password are required")
    save_credentials(body.inventory_id, body.username, body.password)
    return {"message": "Credentials saved"}

@app.get("/api/credentials/{inventory_id}")
async def fetch_credentials(inventory_id: int):
    creds = get_credentials(inventory_id)
    if not creds:
        return {"username": "", "has_credentials": False}
    return {"username": creds["username"], "has_credentials": True, "updated_at": creds["updated_at"]}

# ── host endpoints ────────────────────────────────────────────────────────────

@app.get("/api/hosts")
async def get_hosts_endpoint():
    if not os.path.exists(INVENTORY_PATH):
        return {"hosts": []}
    with open(INVENTORY_PATH, "r") as f:
        lines = f.read().splitlines()
    hosts = [l.strip() for l in lines if l.strip() and not l.startswith("[")]
    return {"hosts": hosts}

@app.post("/api/hosts")
async def update_hosts(body: HostsUpdate):
    if not body.hosts:
        raise HTTPException(status_code=400, detail="Host list cannot be empty")
    content = "[all]\n" + "\n".join(h.strip() for h in body.hosts) + "\n"
    os.makedirs(os.path.dirname(INVENTORY_PATH), exist_ok=True)
    with open(INVENTORY_PATH, "w") as f:
        f.write(content)
    return {"message": f"Inventory updated with {len(body.hosts)} hosts", "hosts": body.hosts}

# ── scan endpoints ────────────────────────────────────────────────────────────

@app.post("/api/scans/trigger")
async def trigger_scan(background_tasks: BackgroundTasks):
    creds = get_active_inventory_credentials()
    if not creds:
        raise HTTPException(
            status_code=400,
            detail="No credentials set for the active inventory. Go to Settings to add credentials."
        )
    already_running = any(v in ("running", "pending") for v in running_scans.values())
    if already_running:
        raise HTTPException(status_code=409, detail="A scan is already in progress")

    scan_id    = str(uuid.uuid4())
    scanned_at = datetime.now(timezone.utc)
    running_scans[scan_id] = "pending"
    background_tasks.add_task(run_and_save, scan_id, scanned_at)
    return {"scan_id": scan_id, "status": "started", "scanned_at": scanned_at.isoformat()}

@app.get("/api/scans/current")
async def current_scan():
    for scan_id, status in reversed(list(running_scans.items())):
        if status in ("running", "pending"):
            return {"scanning": True, "scan_id": scan_id, "status": status}
    return {"scanning": False}

@app.get("/api/scans/latest")
async def latest_scan():
    data = get_latest_scan()
    if not data:
        raise HTTPException(status_code=404, detail="No scans found")
    return data

@app.get("/api/scans/history")
async def scan_history():
    return get_scan_history()

@app.get("/api/scans/{scan_id}/status")
async def get_scan_status(scan_id: str):
    return {"scan_id": scan_id, "status": running_scans.get(scan_id, "unknown")}

@app.get("/api/scans/{scan_id}/failures")
async def get_scan_failures_endpoint(scan_id: str):
    data = get_scan_failures(scan_id)
    if not data:
        raise HTTPException(status_code=404, detail="Scan not found")
    return data

@app.get("/api/scheduler/status")
async def scheduler_status():
    job = scheduler.get_job("auto_scan")
    if not job:
        return {"enabled": False}
    return {
        "enabled": True,
        "next_run": job.next_run_time.isoformat() if job.next_run_time else None,
        "interval_minutes": 10
    }

@app.get("/api/hosts/{hostname}/history")
async def get_host_history_endpoint(hostname: str):
    return get_host_history(hostname)

@app.get("/api/hosts/{hostname}/cves")
async def get_host_cves_endpoint(hostname: str):
    return get_host_cves(hostname)

# ── CVE advisories ────────────────────────────────────────────────────────────

@app.get("/api/cves")
async def get_cves():
    return get_cve_details()

# ── host tags ─────────────────────────────────────────────────────────────────

@app.get("/api/tags")
async def list_all_tags():
    return {"tags": get_all_tags()}

@app.get("/api/hosts/{hostname}/tags")
async def get_host_tags(hostname: str):
    return {"hostname": hostname, "tags": get_tags_for_host(hostname)}

@app.post("/api/hosts/{hostname}/tags")
async def add_host_tag(hostname: str, body: TagAdd):
    tag = body.tag.strip().lower()
    if not tag:
        raise HTTPException(status_code=400, detail="Tag cannot be empty")
    if len(tag) > 32:
        raise HTTPException(status_code=400, detail="Tag too long (max 32 chars)")
    add_tag(hostname, tag)
    return {"hostname": hostname, "tags": get_tags_for_host(hostname)}

@app.delete("/api/hosts/{hostname}/tags/{tag}")
async def remove_host_tag(hostname: str, tag: str):
    remove_tag(hostname, tag)
    return {"hostname": hostname, "tags": get_tags_for_host(hostname)}

# ── notification endpoints ────────────────────────────────────────────────────

@app.get("/api/notifications/settings")
async def get_notification_settings_endpoint():
    s = get_notification_settings()
    if s.get("smtp_password"):
        s["smtp_password"] = "••••••••"
    return s

@app.post("/api/notifications/settings")
async def save_notification_settings_endpoint(body: NotificationSettings):
    existing = get_notification_settings()
    password = body.smtp_password
    if password == "••••••••":
        password = existing.get("smtp_password", "")
    save_notification_settings(
        smtp_host=body.smtp_host,
        smtp_port=body.smtp_port,
        smtp_user=body.smtp_user,
        smtp_password=password,
        smtp_from=body.smtp_from,
        recipients=body.recipients,
        tls_enabled=body.tls_enabled,
    )
    return {"message": "Settings saved"}

@app.post("/api/notifications/test")
async def test_notification(background_tasks: BackgroundTasks):
    settings = get_notification_settings()
    if not settings["smtp_host"]:
        raise HTTPException(status_code=400, detail="SMTP host is not configured")
    if not settings["recipients"]:
        raise HTTPException(status_code=400, detail="No recipients configured")

    def _send_test():
        msg = MIMEMultipart()
        msg["From"]    = settings["smtp_from"] or settings["smtp_user"]
        msg["To"]      = ", ".join(settings["recipients"])
        msg["Date"]    = formatdate(localtime=True)
        msg["Subject"] = "Kernexa — Test Email"
        msg.attach(MIMEText(
            "This is a test email from Kernexa.\n\n"
            "If you received this, your SMTP settings are working correctly.",
            "plain"
        ))

        if settings["tls_enabled"]:
            server = smtplib.SMTP(settings["smtp_host"], settings["smtp_port"], timeout=15)
            server.ehlo()
            server.starttls()
            server.ehlo()
        else:
            server = smtplib.SMTP_SSL(settings["smtp_host"], settings["smtp_port"], timeout=15)

        if settings["smtp_user"] and settings["smtp_password"]:
            server.login(settings["smtp_user"], settings["smtp_password"])

        server.sendmail(msg["From"], settings["recipients"], msg.as_string())
        server.quit()
        print(f"Test email sent to {settings['recipients']}")

    try:
        _send_test()
        return {"message": f"Test email sent to {', '.join(settings['recipients'])}"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ── serve React SPA — MUST BE LAST ───────────────────────────────────────────

if os.path.exists("dist"):
    app.mount("/assets", StaticFiles(directory="dist/assets"), name="assets")

@app.get("/{full_path:path}")
async def serve_spa(full_path: str):
    if full_path.startswith("api/"):
        raise HTTPException(status_code=404, detail="Not found")
    static_path = os.path.join("dist", full_path)
    if full_path and os.path.isfile(static_path):
        return FileResponse(static_path)
    if os.path.exists("dist/index.html"):
        return FileResponse("dist/index.html")
    return {"message": "Patch Scan Platform API"}

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)